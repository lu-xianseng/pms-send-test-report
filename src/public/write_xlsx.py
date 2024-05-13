import os
from enum import Enum
import openpyxl
from config import config
from src.public.log_set import log


class WriteReport:
    def __init__(self, report_file_path):
        # 读入相关excel
        self.wb = openpyxl.load_workbook(config.EXCEL_PATH)
        self.front_cover = self.wb["封面页"]
        self.test_report = self.wb["测试报告"]
        self.test_env = self.wb["测试环境"]
        self.unclosed_bug = self.wb["未关闭Buglist"]
        self.filepath = report_file_path
        log.info(f'文件路径：{self.filepath}')

    @staticmethod
    def __replace_cell_value(wookbook, replacements):
        for i in range(1, wookbook.max_row + 1):
            for j in range(1, wookbook.max_column + 1):
                cell = wookbook.cell(i, j)
                if cell.value:
                    for placeholder, replacement in replacements.items():
                        cell.value = cell.value.replace(placeholder, replacement)

    def update_front_cover(self, app, user, test_time, department, ratify):
        replacements = {
            "{{应用}}": app,
            "{{负责人}}": user,
            "{{日期}}": test_time,
            "{{审核}}": department,
            "{{批准}}": ratify
        }
        self.__replace_cell_value(self.front_cover, replacements)

    def update_test_report(self, report_name, version, user, test_time, test_result, plan, test_task, bug_run,
                           case_run, di, link, risk, bugs, ):
        replacements = {
            "{{项目名称}}": report_name,
            "{{版本}}": version,
            "{{负责人}}": user,
            "{{测试周期}}": test_time,
            "{{测试结果}}": test_result,
            "{{测试策略}}": plan,
            "{{测试单链接}}": test_task,
            "{{BUG/任务转测}}": bug_run,
            "{{用例执行}}": case_run,
            "{{遗留DI}}": di,
            "{{专项链接}}": link,
            "{{风险}}": risk
        }
        self.__replace_cell_value(self.test_report, replacements)

        for index, bug in enumerate(bugs):
            self.test_report.merge_cells(
                start_row=index + 17, start_column=2, end_row=index + 17, end_column=4
            )
            self.test_report.cell(index + 17, 1).value = bug.get("bug_id")
            self.test_report.cell(index + 17, 2).value = bug.get("bug_title")
            self.test_report.cell(index + 17, 5).value = bug.get("severity")
            self.test_report.cell(index + 17, 6).value = bug.get("bug_status")
            self.test_report.cell(index + 17, 7).value = bug.get("bug_type")
            # if int(bug.get("severity")) > 2:
            #     self.test_report.cell(index + 17, 8).value = "回归不通过"
            if bug.get('bug_activeReason'):
                self.test_report.cell(index + 17, 8).value = bug.get("bug_activeReason") + "，回归不通过"

    def update_test_env(self, machine_info):

        try:
            self.test_env.cell(11, 2).value = "\n".join(machine_info[0].get("apps").keys())
            row = column = 3
            for machine in machine_info:
                replacements = {
                    "CPU": machine.get("hardware").get("cpu"),
                    "内存": machine.get("hardware").get("mem"),
                    "显卡": machine.get("hardware").get("gpu"),
                    "硬盘": machine.get("hardware").get("hard"),
                    "网卡": machine.get("hardware").get("net"),
                    "组件": machine.get("software").get("assembly"),
                    "内核": machine.get("software").get("kernel"),
                    "系统": machine.get("software").get("image"),
                    "version": "\n".join(machine.get("apps").values()),
                }
                for value in replacements.values():
                    self.test_env.cell(row, column).value = value
                    row += 1
                row = 3
                column += 1
        except IndexError:
            ...

    def update_unclosed_bug(self, bugs):
        class Header(Enum):
            id = "Bug编号"
            product = "所属产品"
            branch = "所属分支"
            module = "所属模块"
            project = "所属项目"
            story = "相关需求"
            task = "相关任务"
            title = "Bug标题"
            keywords = "关键词"
            severity = "严重程度"
            pri = "优先级"
            type = "Bug类型"
            os = "操作系统"
            browser = "浏览器"
            baseline = "发现基线"
            active = "活动"
            trigger = "触发条件"
            affect = "影响"
            repair = "修复点"
            reprt = "重现概率"
            bugstage = "发现阶段"
            exttype = "类型"
            symbol = "限定符"
            age = "引入阶段"
            source = "来源"
            returnEnvironment = "回归环境"
            steps = "重现步骤"
            status = "Bug状态"
            hangup = "是否挂起"
            deadline = "截止日期"
            activatedCount = "激活次数"
            confirmed = "是否确认"
            mailto = "抄送给"
            openedBy = "由谁创建"
            openedDate = "创建日期"
            openedBuild = "影响版本"
            assignedTo = "指派给"
            assignedDate = "指派日期"
            resolvedBy = "解决者"
            resolution = "集成状态"
            resolvedBuild = "解决版本"
            resolvedDate = "解决日期"
            refusedReson = "拒绝原因"
            actualsolve = "实际解决"
            closedBy = "由谁关闭"
            closedDate = "关闭日期"
            closedReson = "关闭类型"
            duplicateBug = "重复ID"
            linkBug = "相关Bug"
            case = "相关用例"
            lastEditedBy = "最后修改者"
            lastEditedDate = "修改日期"
            testtask = "测试单"
            plan = "所属计划"

        for index, header in enumerate(Header):
            self.unclosed_bug.cell(1, index + 1).value = header.value
        for index, bug in enumerate(bugs):
            for j, header in enumerate(Header):
                self.unclosed_bug.cell(index + 2, j + 1).value = bug.get(header.name)

    def __del__(self):
        self.wb.save(self.filepath)


if __name__ == '__main__':
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    test_env = wb["测试环境"]
    print(test_env.cell(2, 2).value)
